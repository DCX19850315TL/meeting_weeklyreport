#!/usr/bin/env python
#_*_ coding:utf-8 _*_
'''
@auther: tanglei
@contact: tanglei_0315@163.com
@file: excel_code.py
@time: 2019/4/25 16:09
'''
import os
import shutil
from openpyxl import load_workbook
from openpyxl import Workbook

class ExcelHandle(object):

    def __init__(self,excel_path,is_excel):
        self.excel_path = excel_path
        self.is_excel = is_excel

    #读取excel的会议名称和对应的会议号
    def read_excel(self):
        excel_data_list = []
        wb = load_workbook(self.excel_path)
        table = wb.get_active_sheet()
        rows = table.max_row + 1
        for i in range(4,rows):
            meeting_name = table.cell(row=i,column=3).value
            meeting_number = table.cell(row=i,column=8).value
            table_data = {meeting_name:meeting_number}
            excel_data_list.append(table_data)
            table_data = {}
        return excel_data_list
    #将最终格式化后的数据列表写入到excel中
    def set_excel_data(self,excel_file,excel_backup_dir,excel_backup_file,response_list):
        if not os.path.exists(excel_backup_dir):
            os.makedirs(excel_backup_dir)
        if os.path.isfile(excel_file):
            backup_file = os.path.join(excel_backup_dir, excel_backup_file)
            shutil.move(excel_file,backup_file)
        Header = {"A1":"会议号","B1":"会议室名称","C1":"开始时间","D1":"结束时间","E1":"参会方数","F1":"合格率","G1":"参会终端视讯号","H1":"不合格端到端信息"}
        wb = Workbook()
        ws = wb.active
        for k,v in Header.items():
            ws[k] = v
            wb.save(excel_file)
        for item in range(len(response_list)):
            Body = {
                "A%d" % (item + 2):response_list[item]["Meeting_Number"],
                "B%d" % (item + 2):response_list[item]["Meeting_Name"],
                "C%d" % (item + 2):response_list[item]["Start_Time"],
                "D%d" % (item + 2):response_list[item]["End_Time"],
                "E%d" % (item + 2):response_list[item]["Number_Count"],
                "F%d" % (item + 2):response_list[item]["Percent"],
                "G%d" % (item + 2):"\n".join(response_list[item]["Number_List"]),
                "H%d" % (item + 2):"\n".join(response_list[item]["Unqualified_List"])
            }
            for k,v in Body.items():
                ws[k] = v
                wb.save(excel_file)